import tkinter as tk
import sqlite3 as sql
from tkinter import messagebox
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
from datetime import date

ventana = tk.Tk()
ventana.title("Modulo")
ventana.geometry("300x150")

frame = tk.Frame(ventana, padx=10, pady=10)
frame.pack(fill="both", expand=True)

# Frame horizontal para Label y Entry
fila_entry = tk.Frame(frame)
fila_entry.pack(fill="x", pady=(0, 10))  # Ocupa todo el ancho

label = tk.Label(fila_entry, text="Marca:")
label.pack(side="left")

entrada = tk.Entry(fila_entry)
entrada.pack(side="left", fill="x", expand=True)  # Se expande para llenar el espacio

# Frame para los botones
boton_frame = tk.Frame(frame)
boton_frame.pack()

boton_extra_frame = tk.Frame(frame)
boton_extra_frame.pack(pady=(5, 0))  # un poco de espacio hacia arriba

def contar():
    marca=entrada.get()
    marca=marca.upper()
    conn=sql.connect("Anteojos.db")
    cursor = conn.cursor()
    if marca=="":
        cursor.execute("SELECT SUM(Cantidad) FROM Modelos WHERE Cantidad>0")
        resultado=cursor.fetchone()
        resultado=int(resultado[0])
        cursor.execute("SELECT SUM(Cantidad) FROM Modelos WHERE Cantidad>0 AND Marca like '%RX%'")
        resultadoRX=cursor.fetchone()
        resultadoRX=int(resultadoRX[0])
        messagebox.showinfo("Resultado",f"Hay {resultado} anteojos en total, {resultadoRX} de receta y {resultado-resultadoRX} de sol.")
    else:
        cursor.execute("SELECT SUM(Cantidad) FROM Modelos WHERE Marca LIKE ? COLLATE NOCASE AND Cantidad>0", (marca,))
        resultado=cursor.fetchone()
        resultado=int(resultado[0])
        messagebox.showinfo("Resultado",f"Hay {resultado} anteojos de la marca {marca}.")
    return

def exportar():
    marca=entrada.get()
    if marca=="":
        messagebox.showerror("Error","Por favor ingrese una marca para generar la lista de precios")
        return
    conn = sql.connect("Anteojos.db")
    cursor=conn.cursor()
    cursor.execute("""
    SELECT Modelo, Color, Precio 
    FROM Modelos 
    WHERE Marca = ? COLLATE NOCASE AND Cantidad>0
    ORDER BY Modelo ASC
    """, (marca,))

    datos = cursor.fetchall()
    exportar_modelos_a_excel(marca, datos)

def exportar_modelos_a_excel(marca, datos):
    marca_mayus = marca.upper()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Modelos"

    # Fila 1: Marca en mayúsculas y negrita
    fecha_actual = date.today().strftime("%d/%m/%Y")     # Para mostrar en el título
    fecha_para_archivo = date.today().strftime("%d-%m-%Y")  # Para el nombre del archivo

    ws.merge_cells("A1:C1")
    ws["A1"] = f"{marca_mayus} - {fecha_actual}"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Fila 2: Encabezados
    ws.append(["Modelo", "Color", "Precio"])
    for col in range(1, 3):  # Columnas A y B
        letra = get_column_letter(col)
        ws[f"{letra}2"].font = Font(bold=True)
        ws[f"{letra}2"].alignment = Alignment(horizontal="center")

    # Filas 3 en adelante: Datos
    for fila in datos:
        ws.append(fila)

    # Ajustar ancho uniforme para ambas columnas
    ancho_uniforme = 20
    for col in range(1, 3):  # A y B
        letra = get_column_letter(col)
        ws.column_dimensions[letra].width = ancho_uniforme

    # Define un borde delgado para todos los lados
    borde_delgado = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# Aplica el borde a todas las celdas con datos
    for fila in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for celda in fila:
            celda.border = borde_delgado

    # Guardar archivo
    nombre_archivo = f"{marca_mayus}_{fecha_para_archivo}_precios.xlsx"
    wb.save(nombre_archivo)
    messagebox.showinfo("Exportar",f"El archivo se guardo correctamente dentro de la misma carpeta que este programa")


def calcular_lentes():
    lente=entrada.get()
    if lente=="":
        messagebox.showerror("Error","Por favor ingrese un modelo de lente")
        return
    conn = sql.connect("Stock_lentes.db")
    cursor=conn.cursor()
    cursor.execute("""
        SELECT SUM(stock)
        FROM lentes
        WHERE tipo_lente = ? COLLATE NOCASE
        AND columna % 2 = 0
    """, (lente,))

    datos = cursor.fetchone()[0]
    conn.close()

    if datos is None:
        messagebox.showinfo("Resultado", f"Ingreso mal el tipo de lentes. Verifique y vuelva a intentarlo.")
    else:
        messagebox.showinfo("Resultado", f"Hay {datos} lentes del tipo {lente}")

btn_nuevo = tk.Button(boton_extra_frame, text="Calcular lentes", command=calcular_lentes)
btn_nuevo.pack()

btn_guardar = tk.Button(boton_frame, text="Calcular", command=contar)
btn_cancelar = tk.Button(boton_frame, text="Exportar", command=exportar)

btn_guardar.pack(side="left", padx=5)
btn_cancelar.pack(side="left", padx=5)

ventana.mainloop()
